from fastapi import APIRouter, Request, Query
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, Response
from fastapi.templating import Jinja2Templates
from app.dependencies import require_login
from app.database import SessionLocal
from app.models.monitor import Monitor, MonitorStatus
from app.models.monitor_log import MonitorLog, LogStatus
from app.models.incident import Incident
from sqlalchemy import func
from datetime import datetime, timezone, timedelta
import csv
import io

router = APIRouter(prefix="/reports")
templates = Jinja2Templates(directory="app/templates")


def _db_now():
    return datetime.now(timezone.utc).replace(tzinfo=None)


@router.get("/", response_class=HTMLResponse)
async def reports_index(request: Request):
    user = require_login(request)
    db = SessionLocal()
    try:
        monitors = db.query(Monitor).order_by(Monitor.name).all()
        now = _db_now()
        day_ago = now - timedelta(hours=24)
        week_ago = now - timedelta(days=7)
        month_ago = now - timedelta(days=30)

        # --- Uptime summary per monitor (30 days) ---
        uptime_data = []
        for m in monitors:
            total_logs = db.query(func.count(MonitorLog.id)).filter(
                MonitorLog.monitor_id == m.id,
                MonitorLog.checked_at >= month_ago
            ).scalar() or 0
            up_logs = db.query(func.count(MonitorLog.id)).filter(
                MonitorLog.monitor_id == m.id,
                MonitorLog.checked_at >= month_ago,
                MonitorLog.status == LogStatus.up
            ).scalar() or 0
            uptime_pct = (up_logs / total_logs * 100) if total_logs > 0 else None
            avg_rt = db.query(func.avg(MonitorLog.response_time)).filter(
                MonitorLog.monitor_id == m.id,
                MonitorLog.checked_at >= month_ago,
                MonitorLog.status == LogStatus.up
            ).scalar()
            uptime_data.append({
                "monitor": m,
                "total_checks": total_logs,
                "up_checks": up_logs,
                "uptime_pct": uptime_pct,
                "avg_response_time": avg_rt,
            })

        # --- Incident summary ---
        active_incidents = db.query(Incident).filter(Incident.ended_at == None).count()
        total_incidents_30d = db.query(Incident).filter(Incident.started_at >= month_ago).count()
        avg_incident_duration = db.query(func.avg(Incident.duration_seconds)).filter(
            Incident.started_at >= month_ago, Incident.duration_seconds != None
        ).scalar()

        # --- 24h check stats ---
        checks_24h = db.query(func.count(MonitorLog.id)).filter(MonitorLog.checked_at >= day_ago).scalar() or 0
        failures_24h = db.query(func.count(MonitorLog.id)).filter(
            MonitorLog.checked_at >= day_ago, MonitorLog.status == LogStatus.down
        ).scalar() or 0
        checks_7d = db.query(func.count(MonitorLog.id)).filter(MonitorLog.checked_at >= week_ago).scalar() or 0
        failures_7d = db.query(func.count(MonitorLog.id)).filter(
            MonitorLog.checked_at >= week_ago, MonitorLog.status == LogStatus.down
        ).scalar() or 0

        # --- Top slowest monitors ---
        slow_monitors = sorted(
            [u for u in uptime_data if u["avg_response_time"] is not None],
            key=lambda x: x["avg_response_time"], reverse=True
        )[:5]

        # --- Recent incidents (last 30 days) ---
        recent_incidents = db.query(Incident).order_by(
            Incident.started_at.desc()
        ).limit(20).all()

        # --- Daily check counts (last 7 days) ---
        daily_stats = []
        for i in range(6, -1, -1):
            day_start = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
            day_end = day_start + timedelta(days=1)
            total = db.query(func.count(MonitorLog.id)).filter(
                MonitorLog.checked_at >= day_start, MonitorLog.checked_at < day_end
            ).scalar() or 0
            failures = db.query(func.count(MonitorLog.id)).filter(
                MonitorLog.checked_at >= day_start, MonitorLog.checked_at < day_end,
                MonitorLog.status == LogStatus.down
            ).scalar() or 0
            daily_stats.append({
                "date": day_start.strftime("%d.%m"),
                "total": total,
                "failures": failures,
                "success": total - failures,
            })

        return templates.TemplateResponse("reports/index.html", {
            "request": request,
            "user": user,
            "uptime_data": uptime_data,
            "active_incidents": active_incidents,
            "total_incidents_30d": total_incidents_30d,
            "avg_incident_duration": avg_incident_duration,
            "checks_24h": checks_24h,
            "failures_24h": failures_24h,
            "checks_7d": checks_7d,
            "failures_7d": failures_7d,
            "slow_monitors": slow_monitors,
            "recent_incidents": recent_incidents,
            "daily_stats": daily_stats,
        })
    finally:
        db.close()


def _collect_report_data(db):
    """Collect all report data used by export endpoints."""
    now = _db_now()
    month_ago = now - timedelta(days=30)
    monitors = db.query(Monitor).order_by(Monitor.name).all()

    uptime_rows = []
    for m in monitors:
        total_logs = db.query(func.count(MonitorLog.id)).filter(
            MonitorLog.monitor_id == m.id, MonitorLog.checked_at >= month_ago
        ).scalar() or 0
        up_logs = db.query(func.count(MonitorLog.id)).filter(
            MonitorLog.monitor_id == m.id, MonitorLog.checked_at >= month_ago,
            MonitorLog.status == LogStatus.up
        ).scalar() or 0
        uptime_pct = round(up_logs / total_logs * 100, 2) if total_logs > 0 else 0
        avg_rt = db.query(func.avg(MonitorLog.response_time)).filter(
            MonitorLog.monitor_id == m.id, MonitorLog.checked_at >= month_ago,
            MonitorLog.status == LogStatus.up
        ).scalar()
        uptime_rows.append({
            "name": m.name,
            "type": m.type.value,
            "url": m.url,
            "group": m.group.name if m.group else "-",
            "total": total_logs,
            "up": up_logs,
            "uptime": f"{uptime_pct}%",
            "avg_rt": round(avg_rt, 1) if avg_rt else "-",
            "status": m.status.value if m.status else "pending",
        })

    incidents = db.query(Incident).filter(
        Incident.started_at >= month_ago
    ).order_by(Incident.started_at.desc()).all()
    incident_rows = []
    for inc in incidents:
        duration_min = round(inc.duration_seconds / 60, 1) if inc.duration_seconds else "-"
        incident_rows.append({
            "monitor": inc.monitor.name if inc.monitor else "-",
            "start": inc.started_at.strftime("%Y-%m-%d %H:%M"),
            "end": inc.ended_at.strftime("%Y-%m-%d %H:%M") if inc.ended_at else "Devam ediyor",
            "duration": duration_min,
            "reason": inc.reason or "-",
        })

    daily_rows = []
    for i in range(6, -1, -1):
        ds = (now - timedelta(days=i)).replace(hour=0, minute=0, second=0, microsecond=0)
        de = ds + timedelta(days=1)
        total = db.query(func.count(MonitorLog.id)).filter(
            MonitorLog.checked_at >= ds, MonitorLog.checked_at < de
        ).scalar() or 0
        failures = db.query(func.count(MonitorLog.id)).filter(
            MonitorLog.checked_at >= ds, MonitorLog.checked_at < de,
            MonitorLog.status == LogStatus.down
        ).scalar() or 0
        rate = round((total - failures) / total * 100, 1) if total > 0 else 0
        daily_rows.append({
            "date": ds.strftime("%Y-%m-%d"),
            "total": total,
            "success": total - failures,
            "failures": failures,
            "rate": f"{rate}%",
        })

    return now, uptime_rows, incident_rows, daily_rows


UPTIME_HEADERS = ["Monitör", "Tip", "URL", "Grup", "Toplam Kontrol", "Başarılı", "Uptime %", "Ort. Yanıt (ms)", "Durum"]
INCIDENT_HEADERS = ["Monitör", "Başlangıç", "Bitiş", "Süre (dk)", "Sebep"]
DAILY_HEADERS = ["Tarih", "Toplam Kontrol", "Başarılı", "Başarısız", "Başarı Oranı %"]


@router.get("/export")
async def reports_export(request: Request, format: str = Query("csv")):
    user = require_login(request)
    db = SessionLocal()
    try:
        now, uptime_rows, incident_rows, daily_rows = _collect_report_data(db)
        ts = now.strftime('%Y%m%d_%H%M')

        if format == "csv":
            return _export_csv(ts, uptime_rows, incident_rows, daily_rows)
        elif format == "excel":
            return _export_excel(ts, uptime_rows, incident_rows, daily_rows)
        elif format == "pdf":
            return _export_pdf(ts, uptime_rows, incident_rows, daily_rows)
        else:
            return _export_csv(ts, uptime_rows, incident_rows, daily_rows)
    finally:
        db.close()


def _export_csv(ts, uptime_rows, incident_rows, daily_rows):
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    writer.writerow(["=== MONİTÖR UPTIME RAPORU (Son 30 Gun) ==="])
    writer.writerow(UPTIME_HEADERS)
    for r in uptime_rows:
        writer.writerow([r["name"], r["type"], r["url"], r["group"], r["total"], r["up"], r["uptime"], r["avg_rt"], r["status"]])

    writer.writerow([])
    writer.writerow(["=== OLAY GECMİSİ (Son 30 Gun) ==="])
    writer.writerow(INCIDENT_HEADERS)
    for r in incident_rows:
        writer.writerow([r["monitor"], r["start"], r["end"], r["duration"], r["reason"]])

    writer.writerow([])
    writer.writerow(["=== GUNLUK İSTATİSTİKLER (Son 7 Gun) ==="])
    writer.writerow(DAILY_HEADERS)
    for r in daily_rows:
        writer.writerow([r["date"], r["total"], r["success"], r["failures"], r["rate"]])

    # UTF-8 BOM + content for proper Turkish character display in Excel
    bom = b'\xef\xbb\xbf'
    content = bom + output.getvalue().encode('utf-8')
    return Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=rapor_{ts}.csv"}
    )


def _export_excel(ts, uptime_rows, incident_rows, daily_rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    title_font = Font(bold=True, size=13)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    def write_sheet(ws, title, headers, rows, row_keys):
        ws.title = title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value=title).font = title_font
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row_idx, r in enumerate(rows, 3):
            for col_idx, key in enumerate(row_keys, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=r[key])
                cell.border = thin_border
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = 18

    # Sheet 1 - Uptime
    ws1 = wb.active
    write_sheet(ws1, "Uptime Raporu", UPTIME_HEADERS, uptime_rows,
                ["name", "type", "url", "group", "total", "up", "uptime", "avg_rt", "status"])

    # Sheet 2 - Incidents
    ws2 = wb.create_sheet()
    write_sheet(ws2, "Olay Gecmisi", INCIDENT_HEADERS, incident_rows,
                ["monitor", "start", "end", "duration", "reason"])

    # Sheet 3 - Daily
    ws3 = wb.create_sheet()
    write_sheet(ws3, "Gunluk Istatistikler", DAILY_HEADERS, daily_rows,
                ["date", "total", "success", "failures", "rate"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=rapor_{ts}.xlsx"}
    )


def _export_pdf(ts, uptime_rows, incident_rows, daily_rows):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import os

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    elements = []

    # Try to register a font that supports Turkish chars
    font_name = 'Helvetica'
    for fpath in [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        '/usr/share/fonts/TTF/DejaVuSans.ttf',
        'C:/Windows/Fonts/arial.ttf',
    ]:
        if os.path.exists(fpath):
            pdfmetrics.registerFont(TTFont('TurkishFont', fpath))
            font_name = 'TurkishFont'
            break

    title_style = styles['Title']
    title_style.fontName = font_name
    heading_style = styles['Heading2']
    heading_style.fontName = font_name

    elements.append(Paragraph("FBU Downtime Monitor - Rapor", title_style))
    elements.append(Spacer(1, 5*mm))

    header_bg = colors.HexColor('#2563EB')
    header_text = colors.white
    table_style_base = [
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), header_text),
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]

    # Uptime table
    elements.append(Paragraph("Monitor Uptime Raporu (Son 30 Gun)", heading_style))
    elements.append(Spacer(1, 3*mm))
    data = [UPTIME_HEADERS]
    for r in uptime_rows:
        data.append([r["name"], r["type"], r["url"], r["group"], r["total"], r["up"], r["uptime"], r["avg_rt"], r["status"]])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle(table_style_base))
    elements.append(t)
    elements.append(Spacer(1, 8*mm))

    # Incidents table
    elements.append(Paragraph("Olay Gecmisi (Son 30 Gun)", heading_style))
    elements.append(Spacer(1, 3*mm))
    data = [INCIDENT_HEADERS]
    for r in incident_rows:
        data.append([r["monitor"], r["start"], r["end"], r["duration"], r["reason"]])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle(table_style_base))
    elements.append(t)
    elements.append(Spacer(1, 8*mm))

    # Daily table
    elements.append(Paragraph("Gunluk Istatistikler (Son 7 Gun)", heading_style))
    elements.append(Spacer(1, 3*mm))
    data = [DAILY_HEADERS]
    for r in daily_rows:
        data.append([r["date"], r["total"], r["success"], r["failures"], r["rate"]])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle(table_style_base))
    elements.append(t)

    doc.build(elements)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=rapor_{ts}.pdf"}
    )
